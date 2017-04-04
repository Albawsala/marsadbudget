# -*- coding: utf-8 -*-

from werkzeug.exceptions import BadRequest, NotFound
from flask import render_template, request, url_for, flash, redirect, current_app, g, abort
from flask_login import login_user, current_user, login_required, logout_user
from openpyxl import load_workbook
from os import remove
from helpers import *
from common.users import *
#abs
from bson.dbref import DBRef
from datetime import datetime

import collections as collections_
import documents as documents_
import files as files_
import json
import re
#abs
import os
import iso8601
from pprint import pprint
ALLOWED_EXTENSIONS = set(['xlsx'])
UPLOAD_FOLDER = '/home/user0/www/uploads/marsad_budget/documents/'
def test():
	return 'Hello World'


def index():
	if current_user.is_authenticated():
		return render_template('index.html')
	return redirect(url_for('admin.login'))


@login_required
@editor_required
def collections():
	query = {}
	if current_user.data["role"] == "editor" and current_user.data.get("limited"):
		query['_name'] = {'$in': current_user.data["collections"]}
	collections = g.db.find('collections', query)
	return render_template('collections.html',
		collections = collections,
	)


@login_required
@admin_required
def add_collection():
	_csrf_key = 'csrf_collections_add'
	if request.method == 'POST':
		if not request.json and not check_csrf_token(request.json, _csrf_key):
			raise BadRequest("Request check failed")
		try:
			request.json.pop("_csrf")
			collections_.save_collection(request.json)
			return 'Collection Added Successfully!'

		except AssertionError as err:
			return u'{"message": "<strong>Error: </strong>%s", "csrf": "%s"}' % (str(err), generate_csrf_token(_csrf_key)), 400

	return render_template('collection_add.html',
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
@admin_required
def del_collection(collection):
	_csrf_key = 'csrf_collections_del_%s' % collection
	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			collections_.remove_collection(collection)
			return redirect(url_for('admin.collections'))
		except AssertionError as err:
			raise BadRequest(err)

	return render_template('delete.html',
		obj_type = 'collection',
		obj_url = url_for('admin.del_collection', collection=collection),
		csrf = generate_csrf_token(_csrf_key),
	)


@login_required
@admin_required
def edit_collection(collection):
	col = g.db.find_one('collections', {'_name': collection})
	if not collection:
		raise NotFound()

	_csrf_key = 'csrf_collections_edit_%s' % collection

	if request.method == 'POST':
		if not request.json and not check_csrf_token(request.json, _csrf_key):
			raise BadRequest("Request check failed")
		try:
			request.json.pop("_csrf")
			collections_.update_collection(col, request.json)
			return 'Collection Updated Successfully!'

		except AssertionError as err:
			return u'{"message": "<strong>Error: </strong>%s", "csrf": "%s"}' % (str(err), generate_csrf_token(_csrf_key)), 400

	return render_template('collection_edit.html',
		collection = col,
		csrf = generate_csrf_token(_csrf_key),
	)


@login_required
@editor_required
def documents(collection):
	col = g.db.find_one('collections', {'_name': collection}) or abort(404)

	docs, first_p, last_p = setup_pagination(collection, current_app.config['PAGE_MAX'])
	featured = collections_.get_featured(col)
	fields = collections_.get_fields_paths(col["_schema"])

	if request.is_xhr:
		template = 'documents_ajax.html'
	else:
		template = 'documents.html'

	return render_template(template,
		collection = collection,
		docs = docs,
		featured = featured,
		fields = fields,
		first_p = first_p,
		last_p = last_p,
	)


@login_required
@editor_required
def add_document(collection, doc_id):
	col = g.db.find_one('collections', {'_name': collection}) or abort(404)

	_csrf_key = 'csrf_%s_add' % collection

	if request.method == 'POST':
		if not request.json and not check_csrf_token(request.json, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			doc_new = request.json.get("doc")
			is_draft = request.json.get("draft")
			is_iframe = request.args.get("iframe")
			assert doc_new, "document not found"
			assert isinstance(is_draft, bool), "document meta info missing"
			_id = documents_.insert_document(col, doc_new, is_draft)
			doc_id = str(_id)
			return json.dumps({
				'message': 'Document Added Successfully!',
				'_id': doc_id,
				'edit_url': url_for('admin.edit_document', collection=collection, doc_id=doc_id),
				'featured': u''.join(map(lambda getter: u'<span>%s</span>' % getter(doc_new), collections_.get_featured(col)[1])),
			})

		except AssertionError as e:
			return json.dumps({
				"message": "<strong>Error: </strong>%s" % str(e),
				"csrf": generate_csrf_token(_csrf_key)
			}), 400

	if request.is_xhr:
		return json.dumps({'csrf': generate_csrf_token(_csrf_key)})

	doc = None
	if doc_id:
		doc = g.db.find_one(collection, {'_id': doc_id}) or abort(404)

	return render_template('document_add.html',
		collection = col,
		doc = doc,
		csrf = generate_csrf_token(_csrf_key),
	)


@login_required
@editor_required
def del_document(collection, doc_id):
	col = g.db.find_one('collections', {'_name': collection}) or abort(404)
	doc = g.db.find_one(collection, {'_id': doc_id}) or abort(404)

	_csrf_key = 'csrf_%s_del_%s' % (collection, doc_id)

	if request.method == 'POST':
		if current_user.data.get("limited") and collection not in current_user.data.get("collections"):
			raise BadRequest("Access denied")
		if request.is_xhr:
			if not check_csrf_token(request.json, _csrf_key):
				raise BadRequest("Request check failed")
			documents_.remove_document(collection, doc)
			return json.dumps({'message': 'Document Deleted Successfully!'})
		else:
			if not check_csrf_token(request.form, _csrf_key):
				raise BadRequest("Request check failed")
			documents_.remove_document(collection, doc)
			return redirect(url_for('admin.documents', collection=collection))

	if request.is_xhr:
		return json.dumps({'csrf': generate_csrf_token(_csrf_key)})

	return render_template('delete.html',
		obj_type = 'document',
		obj_url = url_for('admin.del_document', collection=collection, doc_id=doc_id),
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
@editor_required
def edit_document(collection, doc_id):
	col = g.db.find_one('collections', {'_name': collection}) or abort(404)
	doc = g.db.find_one(collection, {'_id': doc_id}) or abort(404)

	_csrf_key = 'csrf_%s_edit_%s' % (collection, doc_id)

	if request.method == 'POST':
		if current_user.data.get("limited") and collection not in current_user.data.get("collections"):
			raise BadRequest("Access denied")
		if not request.json and not check_csrf_token(request.json, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			doc_new = request.json.get("doc")
			is_draft = request.json.get("draft")
			mtime = request.json.get("mtime")
			assert doc_new, "document not found"
			assert isinstance(is_draft, bool), "document meta info missing"
			documents_.update_document(col, doc, doc_new, is_draft, mtime)
			return json.dumps({
				'message': 'Document Updated Successfully!',
				'_id': doc_id,
				'edit_url': url_for('admin.edit_document', collection=collection, doc_id=doc_id),
				'featured': u''.join(map(lambda getter: u'<span>%s</span>' % getter(doc_new), collections_.get_featured(col)[1])),
			})

		except AssertionError as err:
			return json.dumps({
				"message": "<strong>Error: </strong>%s" % str(err),
				"csrf": generate_csrf_token(_csrf_key),
			}), 400

	is_open = False
	now = datetime.now()
	me = str(current_user.get_id())
	for toucher, touched in doc['_meta'].setdefault('touch', {}).items():
		if toucher != me and (now - touched).seconds < 90:
			is_open = True
			break

	documents_.touch_document(collection, doc_id)

	if request.is_xhr:
		return json.dumps({
			'csrf': generate_csrf_token(_csrf_key),
			'mtime': doc['_meta']['mtime'].isoformat(),
		})

	return render_template('document_edit.html',
		collection = col,
		doc = doc,
		is_open = is_open,
		csrf = generate_csrf_token(_csrf_key),
	)


@login_required
@editor_required
def touch_document(collection, doc_id):
	if request.is_xhr:
		documents_.touch_document(collection, doc_id)
	return "ping", 200


@login_required
@editor_required
def find_document(collection):
	col = g.db.find_one('collections', {'_name': collection}) or abort(404)

	query = request.form.get('_query')
	path = request.form.get('_path')
	field = request.form.get(path)
	query and path and field or abort(404)

	path = re.sub('^root\.?', '', path)
	docs = list(documents_.find_documents(col, query, path, field))
	featured = collections_.get_featured(col)
	fields = collections_.get_fields_paths(col["_schema"])

	if request.is_xhr:
		template = 'documents_ajax.html'
	else:
		template = 'documents.html'

	return render_template(template,
		collection = collection,
		featured = featured,
		fields = fields,
		docs = docs,
		first_p = True,
		last_p = True,
	)


FILE_FIELDS = [
	{'name': 'title', 'type': 'string'},
	{'name': 'filename', 'type': 'string'},
	{'name': 'type', 'type': 'choice'},
	{'name': 'lang', 'type': 'choice'},
]


@login_required
@editor_required
def files():
	liste, first_p, last_p = setup_pagination('files', current_app.config['FILES_PAGE_MAX'])

	if request.is_xhr:
		template = 'files_ajax.html'
	else:
		template = 'files.html'

	return render_template(template,
		files = liste,
		fields = FILE_FIELDS,
		first_p = first_p,
		last_p = last_p,
	)


@login_required
@editor_required
def add_file():
	_csrf_key = 'csrf_files_add'

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			fs = request.files.get("file")
			type_ = request.form.get("type")
			title_ = request.form.get("title")
			lang_ = request.form.get("lang")
			assert fs, "missing file"
			assert type_ and type_ in ["DOC","IMG"], "missing or invalid type"
			assert title_, "missing title"
			if type_ == "DOC":
				assert lang_ in ["fr","ar","en"], "missing on invalid lang"
			elif type_ == "IMG":
				assert "width" in request.form, "missing width"

			_id = files_.insert_file(request.form, fs)
			is_iframe = request.args.get('iframe')
			if is_iframe:
				return render_template('file_update.html',
					iframe = is_iframe,
					file_id = _id,
					edit_url = url_for('admin.edit_file', file_id=_id),
					featured = u'<span>%s</span><span>%s</span>' % (title_, type_),
				)
			flash('File Added Successfully!')

		except AssertionError as err:
			flash(err)

	return render_template('file_add.html',
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
@editor_required
def del_file(file_id):
	_csrf_key = 'csrf_files_del_%s' % file_id

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		files_.remove_file(file_id)
		return redirect(url_for('admin.files'))

	return render_template('delete.html',
		obj_type = 'file',
		obj_url = url_for('admin.del_file', file_id=file_id),
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
@editor_required
def edit_file(file_id):
	file_ = g.db.find_one('files', {'_id':file_id})
	if not file_:
		raise NotFound()

	_csrf_key = 'csrf_files_edit_%s' % file_id

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			title_ = request.form.get("title")
			lang_ = request.form.get("lang")
			assert title_, "title missing"
			if file_['type'] == "DOC":
				assert lang_ in ["fr","ar","en"], "missing on invalid lang"
			elif file_['type'] == "IMG":
				assert "width" in request.form, "width missing"

			files_.update_file(request.form, file_)
			is_iframe = request.args.get('iframe')
			if is_iframe:
				return render_template('file_update.html',
					iframe = is_iframe,
					file_id = file_id,
					edit_url = url_for('admin.edit_file', file_id=file_id),
					featured = u'<span>%s</span><span>%s</span>' % (title_, file_['type']),
				)
			flash('File Updated Successfully!')

		except AssertionError as err:
			flash(err)

	return render_template('file_edit.html',
		csrf = generate_csrf_token(_csrf_key),
		file = file_
	)


@login_required
@editor_required
def find_file():
	query = request.form.get('_query')
	path = request.form.get('_path')
	field = request.form.get(path)
	query and path and field or abort(404)

	files = list(files_.find_files(field, query))

	if request.is_xhr:
		template = 'files_ajax.html'
	else:
		template = 'files.html'

	return render_template(template,
		files = files,
		fields = FILE_FIELDS,
		first_p = True,
		last_p = True,
	)


@login_required
@editor_required
def images():
	return render_template('images.html',
		liste = g.db.find('files', {'type': 'IMG'}, 11)
	)


@login_required
@editor_required
def add_image():
	_csrf_key = 'csrf_images_add'

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			fs = request.files.get('file')
			assert fs, 'upload failed'
			assert all(k in request.form for k in ["title", "width", "type"]), "missing fields"
			filepath = files_.insert_file(request.form, fs, mce_image=True)
			return redirect(url_for('admin.images', editor=request.args.get('editor')))

		except AssertionError as err:
			raise BadRequest(str(err))

	return render_template('image_add.html',
		csrf = generate_csrf_token(_csrf_key)
	)


def login():
	if current_user.is_authenticated():
		return redirect(request.args.get('next') or '/')

	_csrf_key = 'csrf_user_login'

	if request.method == 'POST':

		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			email = request.form.get('email')
			password = request.form.get('password')
			remember = request.form.get('remember')
			assert email, "missing email"
			assert password, "missing password"

			user = get_user(email, password)
			assert user, "wrong email or password"

			login_user(user, remember=remember)
			flash("Logged in successfully.")
			return redirect(request.args.get("next") or url_for('admin.index'))

		except AssertionError as err:
			flash(err)

	return render_template("login.html",
		csrf = generate_csrf_token(_csrf_key)
	)

@login_required
@admin_required
def register():
	_csrf_key = 'csrf_user_register'

	if request.method == "POST":
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			form_name = request.form.get('name')
			form_email = request.form.get('email')
			form_password = request.form.get('password')
			form_confirm = request.form.get('confirm')

			assert form_name, "missing name"
			assert form_email, "missing email"
			assert form_password, "missing password"
			assert form_password == form_confirm, "password confirm wrong"

			add_user(form_name, form_email, form_password)
			flash("User Registred Successfully!")
			return redirect(url_for('admin.index'))

		except AssertionError as err:
			flash(err)

	return render_template("register.html",
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
def logout():
	logout_user()
	return redirect(url_for("admin.index"))


@login_required
@editor_required
def profile():
	_csrf_key = 'csrf_user_edit'

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			form_name = request.form.get('name')
			form_email = request.form.get('email')
			form_password = request.form.get('password')
			form_confirm = request.form.get('confirm')

			assert form_name, "missing name"
			assert form_email, "missing email"
			assert form_password, "missing password"
			assert form_password == form_confirm, "password confirm wrong"

			update_profile(current_user.data, form_name, form_email, form_password)
			flash("Profile Updated Successfully!")

		except AssertionError as err:
			flash(err)

	return render_template('profile.html',
		csrf = generate_csrf_token(_csrf_key)
	)


@login_required
@editor_required
def add_collection_shortcut(collection):
	add_shortcut(current_user, collection)
	return redirect(request.args.get("next") or url_for('admin.collections'))


@login_required
@editor_required
def del_collection_shortcut(collection):
	del_shortcut(current_user, collection)
	return redirect(request.args.get("next") or url_for('admin.collections'))


@login_required
@admin_required
def users():
	return render_template('users.html',
		users = g.db.find('users')
	)

@login_required
@admin_required
def user_edit(user_id):
	user = g.db.find_one('users', {'_id':user_id})
	if not user:
		raise NotFound()

	_csrf_key = 'csrf_users_edit_%s' % user_id

	if request.method == 'POST':
		if not check_csrf_token(request.form, _csrf_key):
			raise BadRequest("Request check failed")

		try:
			assert all(name in request.form for name in ["name", "email", "role", "collections"]), "form invalid"
			assert request.form.get("role") in ['subscriber', 'editor', 'admin'], 'role field invalid'

			update_user(user, request.form)
			flash('User Updated Successfully!')

		except AssertionError as err:
			flash(err)

	return render_template('user_edit.html',
		user = user,
		csrf = generate_csrf_token(_csrf_key)
	)

def Project_XL_JSON(filename):
	wb = load_workbook(filename)
	projets = []
	sheet_fr = wb.get_sheet_by_name(wb.get_sheet_names()[0]) # first sheet should be in french
	sheet_ar = wb.get_sheet_by_name(wb.get_sheet_names()[1]) # second sheet shoud be in arabic
	for row_fr , row_ar in zip(sheet_fr.rows , sheet_ar.rows):
	    projet = {
	    'titre'         : { 'fr' : row_fr[0].value.encode('utf8') , 'ar' :
row_ar[0].value.encode('utf8')},
	    'ministere'     : { 'fr' : row_fr[1].value.encode('utf8') , 'ar' :
row_ar[1].value.encode('utf8')},
	    'secteur'       : { 'fr' : row_fr[2].value.encode('utf8') , 'ar' :
row_ar[2].value.encode('utf8')},
	    'cout_prev'     : row_fr[3].value ,
	    'cout_reel'     : row_fr[4].value ,
	    'date_deb'      : row_fr[5].value ,
	    'date_fin'      : row_fr[6].value ,
	    'date_fin_reel' : row_fr[7].value ,
	    'contexte'      : { 'fr' : row_fr[8].value.encode('utf8') , 'ar' :
row_ar[8].value.encode('utf8')},
	    'contenu'       : { 'fr' : row_fr[9].value.encode('utf8') , 'ar' :
row_ar[9].value.encode('utf8')},
	    'objectif'      : { 'fr' : row_fr[10].value.encode('utf8'), 'ar' :
row_ar[10].value.encode('utf8')},
	    'financement'   : { 'fr' : row_fr[11].value.encode('utf8'), 'ar' :
row_ar[11].value.encode('utf8')},
	    'coord'         : { 'x'  : row_fr[12].value , 'y' : row_fr[13].value  },
	    'delegation'    : { 'fr' : row_fr[14].value.encode('utf8'), 'ar' :
row_ar[14].value.encode('utf8')},
	    'municipalite'  : { 'fr' : row_fr[15].value.encode('utf8'), 'ar' :
row_ar[15].value.encode('utf8')},
	    'gouvernorat'   : { 'fr' : row_fr[16].value.encode('utf8'), 'ar' :
row_ar[16].value.encode('utf8')},
		'progress'		: row_fr[17].value ,
	    }
	    projets.append(projet)
	return projets

@login_required
@editor_required
def add_project():
	UPLOAD_FOLDER = '/home/user0/www/uploads/marsad_budget/documents/'
	ALLOWED_EXTENSIONS = set(['xlsx'])
	if request.method == "POST":
		if 'project_XL' not in request.files:
			return render_template('add_project.html',error ='select file please')
		else:
			xl = request.files['project_XL']
			if xl.filename == '':
				return render_template('add_project.html',error ='file not allowed')
			else:
				if xl.filename[-4:] in ALLOWED_EXTENSIONS:
					xl.save(UPLOAD_FOLDER+xl.filename)
					for projet in Project_XL_JSON(UPLOAD_FOLDER+xl.filename):
						if g.db.find("secteurs",{ 'nom.fr'  :
projet['secteur']['fr']}).count() == 0:
							g.db.insert('secteurs',{'nom' : { 'fr' :
projet['secteur']['fr'] , 'ar' : projet['secteur']['ar']}})
						if g.db.find("financement",{ 'nom.fr'  :
projet['secteur']['fr']}).count() == 0:
							g.db.insert('financement',{'nom' : { 'fr' :
projet['financement']['fr'] , 'ar' : projet['financement']['ar']}})
						if g.db.find("projects",{'titre.fr' :
projet['titre']['fr']}).count() == 0:
							g.db.insert('projects',projet)
						else:
							g.db.update('projects', {'titre.fr' :
projet['titre']['fr']} , projet)
					remove(UPLOAD_FOLDER+xl.filename)
				else:
					return render_template('add_project.html',error ="file extension not allowed")
		return render_template('add_project.html',error ="success project added")
	else:
		return render_template('add_project.html')


@login_required
@editor_required
def view_projects():
	projects = g.db.find("projects",)
	return render_template("projects.html",projects=projects)

@login_required
@editor_required
def edit_project(project_id):
	project = g.db.find_one("projects" , {'_id' : project_id }) or abort(404)
	if request.method == "GET":
		return render_template('edit_project.html' , project = project)
	else:
		edited_project = {
	    'titre'         : { 'fr' : request.form.get('titre_fr') 	, 'ar' :
request.form.get('titre_ar')},
	    'ministere'     : { 'fr' : request.form.get('minis_fr') 	, 'ar' :
request.form.get('minis_ar')},
	    'secteur'       : { 'fr' : request.form.get('secteur_fr') 	, 'ar' :
request.form.get('secteur_ar')},
	    'cout_prev'     : request.form.get('cout_prev'),
	    'cout_reel'     : request.form.get('cout_reel'),
	    'date_deb'      : request.form.get('date_deb') ,
	    'date_fin'      : request.form.get('date_fin') ,
	    'date_fin_reel' : request.form.get('date_fin_reel'),
	    'contexte'      : { 'fr' : request.form.get('contexte_fr') 	, 'ar' :
request.form.get('contexte_ar')},
	    'contenu'       : { 'fr' : request.form.get('contenu_fr') 	, 'ar' :
request.form.get('contenu_ar')},
	    'objectif'      : { 'fr' : request.form.get('objectif_fr') 	, 'ar' :
request.form.get('objectif_ar')},
	    'financement'   : { 'fr' : request.form.get('financement_fr') ,'ar':
request.form.get('financement_ar')},
	    'coord'         : { 'x'  : request.form.get('coord_x') 		, 'y'  :
request.form.get('coord_y')  },
	    'delegation'    : { 'fr' : request.form.get('delegation_fr'), 'ar' :
request.form.get('delegation_ar')},
	    'municipalite'  : { 'fr' : request.form.get('municipalite_fr'), 'ar' :
request.form.get('municipalite_ar')},
	    'gouvernorat'   : { 'fr' : request.form.get('gov_fr')	, 'ar' :
request.form.get('gov_ar')},
		'progress'		: request.form.get('progress') ,
	    }
		g.db.update('projects', {'_id' : project["_id"]} , edited_project)
		return redirect(url_for('admin.projects'))
@login_required
@editor_required
def delete_project(project_id):
	project = g.db.find_one("projects" , {'_id' : project_id }) or abort(404)
	g.db.remove("projects" , { '_id' : project["_id"] })
	return redirect(url_for('admin.projects'))


@login_required
@editor_required
def show_resources():
	resources= g.db.find('resources',)
	return render_template('resources.html',resources=resources)

def resources_to_json(filename):
	wb = load_workbook(filename)
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	resources = {}
	for row in sheet.rows:
	    val = {
		'children'  : [],
		'total'   : row[3].value ,
	    'name'      : { 'fr' : row[2].value , 'ar' : row[1].value } ,
		'id'        : row[0].value.__str__() ,
	    }
	    name = row[0].value.__str__()
	    if name.__len__() == 1:
	        resources[name] = val
	    elif name.__len__() == 3:
	        resources[name[0]]['children'].append(val)
	    elif name.__len__() == 5:
	        resources[name[0]]['children'][-1]['children'].append(val)
	    elif name.__len__() == 7:
	        resources[name[0]]['children'][-1]['children'][-1]['children'].append(val)
	    elif name.__len__() == 9:
	        resources[name[0]]['children'][-1]['children'][-1]['children'][-1]['children'].append(val)
	return resources

@login_required
@editor_required
def add_resources():
	if request.method == 'POST':
		if 'resources_XL' not in request.files:
			return render_template('add_resource.html',error ='please select a file')
		elif request.form.get('resource_year') == '':
			return render_template('add_resource.html',error ='please type in the year')
		elif request.form.get('resource_type') not in ['LF' , 'LFC']:
			return render_template('add_resource.html',error ='please select a valid type ( LF or LFC)')
		else:
			xl = request.files['resources_XL']
			if xl.filename == '':
				return render_template('add_resource.html',error ='select a file please')
			else:
				if xl.filename[-4:] in ALLOWED_EXTENSIONS:
					xl.save(UPLOAD_FOLDER+xl.filename)
					res = {}
					res['year'] = request.form.get('resource_year')
					res['type'] = request.form.get('resource_type')
					res['data'] = resources_to_json(UPLOAD_FOLDER+xl.filename)
					if g.db.find('resources',{'year' : res['year'].__str__(), 'type' : res['type']}).count() == 0:
						g.db.insert('resources',res)
					else:
						g.db.update('resources',{'year' : res['year'], 'type' : res['type'] },res)
					return redirect(url_for('admin.resources'))
				else:
					return render_template('add_resource.html',error ='select a valid file please')
	return render_template('add_resource.html')

@login_required
@editor_required
def edit_resource(resource_id):
	resource = g.db.find_one("resources" , {'_id' : resource_id }) or abort(404)
	if request.method == "GET":
		return render_template('edit_resource.html' , resource = resource)
	else:
		tp = request.form.get('resource_type')
		yr = request.form.get('resource_year')
		if tp in ['LF','LFC']:
			resource['type'] = tp
			resource['year'] = yr
			g.db.update('resources', {'_id' : resource["_id"]} , resource)
		return redirect(url_for('admin.resources'))


@login_required
@editor_required
def delete_resource(resource_id):
	resource = g.db.find_one("resources" , {'_id' : resource_id }) or abort(404)
	g.db.remove("resources" , { '_id' : resource["_id"] })
	return redirect(url_for('admin.resources'))


#-------------------------------------------------------article affichage------------------------------------------------------ok
@login_required
@editor_required
def show_articles():
	article =[]
	articles = g.db.find('article',sort=[('date',-1)])
	for post in articles:
		post['contenu']['fr']=post['contenu']['fr'][0:150]
		post['titre']['fr']=post['titre']['fr']
		article.append(post)
	return render_template('article.html',articles = article)


#-------------------------------------------------------article insertion------------------------------------------------------ok
@login_required
@editor_required
def new_articles():
	UPLOAD_FOLDER = '/home/user0/www/uploads/marsad_budget/images/'
	ALLOWED_EXTENSIONS = set(['jpeg','.jpg','.png'])
	if request.method == 'POST':
		if request.files['image'].filename[-4:] not in ALLOWED_EXTENSIONS:
			return render_template('add_article.html',
				error = "type de fichier invalide (jpg,jpeg,png)",
				titre_fr = request.form.get('titre_fr'),
				titre_ar = request.form.get('titre_ar'),
				contenu_fr = request.form.get('contenu_fr'),
				contenu_ar=request.form.get('contenu_ar'),)

		else:
#			return str(iso8601.parse_date(str(datetime.now())))
			request.files['image'].save(UPLOAD_FOLDER+request.files['image'].filename)
			files = {	"filename" : request.files['image'].filename,
					"lang" : "fr",
					"type" : "IMG",
					"filesize" : os.stat(UPLOAD_FOLDER+request.files['image'].filename).st_size,
					"title" : request.files['image'].filename[:-4],
				}
			g.db.insert('files',files)
			article = {
				'titre'	:	{'fr'	:	request.form.get('titre_fr'), 'ar'	:	request.form.get('titre_ar')},
				'contenu' :	{'fr'	:	request.form.get('contenu_fr'),'ar'	:	request.form.get('contenu_ar')},
				'image' :	DBRef("files", str(files['_id'])),
				'date'	:	iso8601.parse_date(str(datetime.now())),
				}
			g.db.insert('article',article)
			return redirect(url_for('admin.articles'))

	else:
		return render_template('add_article.html',titre_fr = "",titre_ar = "", contenu_fr = "",contenu_ar="")

#-------------------------------------------------------article modification------------------------------------------------------ok
@login_required
@editor_required
def edit_articles(article_id):
	article = g.db.find_one("article" , {'_id' : article_id }) or abort(404)
	image = g.db.find_one("files",{"_id" : article['image'].id})
#	return render_template('edit_article.html', article = article)
#	return str(article['image'].id)
	if request.method == "GET":
		return render_template('edit_article.html' , article = article, image = image)
	else:
		UPLOAD_FOLDER = '/home/user0/www/uploads/marsad_budget/images/'
		ALLOWED_EXTENSIONS = set(['jpeg','.jpg','.png'])
		if 'image' not in request.files or request.files['image'].filename=="":
			article = {
				'titre'	:	{'fr'	:	request.form.get('titre_fr'), 'ar'	:	request.form.get('titre_ar')},
				'contenu' :	{'fr'	:	request.form.get('contenu_fr'),'ar'	:	request.form.get('contenu_ar')},
				}
			g.db.update("article",{'_id' : article_id },{'$set':article})
			return render_template('edit_article.html',article=article, error = "modification avec succee")
		elif request.files['image'].filename[-4:] not in ALLOWED_EXTENSIONS:
			return render_template('edit_article.html',article=article, error = "type de fichier invalide (jpg,jpeg,png)")

		else:
			request.files['image'].save(UPLOAD_FOLDER+request.files['image'].filename)
			g.db.update('files',{'_id':article['image'].id},{'$set':{"filename" : request.files['image'].filename,"title" : request.files['image'].filename[:-4],"filesize" : os.stat(UPLOAD_FOLDER+request.files['image'].filename).st_size,}})
			article = {
				'titre'	:	{'fr'	:	request.form.get('titre_fr'), 'ar'	:	request.form.get('titre_ar')},
				'contenu' :	{'fr'	:	request.form.get('contenu_fr'),'ar'	:	request.form.get('contenu_ar')},
				}
			g.db.update("article",{'_id' : article_id },{'$set':article})
			return render_template('edit_article.html',article=article, error = "modification avec succee")

@login_required
@editor_required
def delete_articles(article_id):
	article = g.db.find_one("article" , {'_id' : article_id }) or abort(404)
        g.db.remove("article" , { '_id' : article["_id"] })
        return redirect(url_for('admin.articles'))

def budget_to_json(filename):
	To_return = []
	tmp_dict = {}
	wb = load_workbook(filename ,  data_only=True, read_only=True)
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	data = enumerate(list(sheet.rows)[2:])
	for i,row in data:
		if ( row[0].value and  row[1].value and row[2].value ) :
			tmp_dict["ar"] = row[0].value
			tmp_dict["fr"] = row[1].value
			tmp_dict["total"] = row[2].value
			tmp_dict["remunerations_publique"] =  row[3].value or 0
			tmp_dict["moyens_des_services"] =  row[4].value or 0
			tmp_dict["interventions_publiques"] =  row[5].value or 0
			tmp_dict["disposition_urgence"] =  row[6].value or 0
			tmp_dict["Avantages_dette_publique"] =  row[7].value or 0
			tmp_dict["Investissements_direct"] =  row[9].value or 0
			tmp_dict["financement_public"] =  row[10].value or 0
			tmp_dict["Depenses_developpement_urgence"] =  row[11].value or 0
			tmp_dict["ressources_exterieures_affectees"] =  row[12].value or 0
			tmp_dict["Remboursement_dette_publique"] =  row[13].value or 0
			tmp_dict["fonds_speciaux"] = row[15].value or 0
			tmp_dict["gestion"] =  row[8].value or 0
			tmp_dict["developpement"] =  row[14].value or 0

			To_return.append(tmp_dict)
			tmp_dict = {}
	return To_return

def link_budgets(budgets , year , budget_type):
	for budget in budgets:
		obj = {}
		ministere = g.db.find_one('ministeres' , {'nom.ar' : {'$regex' : budget['ar'] , "$options": 'i'}})
		obj['budget']	 = {}
		obj['budget']['children'] = {}
		obj['budget']['total']	 = budget['total']
		obj['annee']	 = int(year)
		obj['type'] 	 = budget_type
		if budget['developpement'] != 0:
			obj['budget']['children']['developpement'] = {'montant' : budget['developpement'] / 1000000.0 , 'children' : []}
			if budget['Investissements_direct'] != 0:
				obj['budget']['children']['developpement']['children'].append({ 'ref' : 'Investissements_direct' , 'montant' : budget['Investissements_direct'] / 1000000.0})
			if budget['financement_public'] != 0:
				obj['budget']['children']['developpement']['children'].append({ 'ref' : 'financement_public' , 'montant' : budget['financement_public'] / 1000000.0})
			if budget['ressources_exterieures_affectees'] != 0:
				obj['budget']['children']['developpement']['children'].append({ 'ref' : 'ressources_exterieures_affectees' , 'montant' : budget['ressources_exterieures_affectees'] / 1000000.0})
			if budget['Depenses_developpement_urgence'] != 0:
				obj['budget']['children']['developpement']['children'].append({ 'ref' : 'Depenses_developpement_urgence' , 'montant' : budget['Depenses_developpement_urgence'] / 1000000.0})
			if budget['Remboursement_dette_publique'] != 0:
				obj['budget']['children']['developpement']['children'].append({ 'ref' : 'Remboursement_dette_publique' , 'montant' : budget['Remboursement_dette_publique'] / 1000000.0})
		if budget['gestion'] != 0:
			obj['budget']['children']['gestion'] = { 'montant' : budget['gestion'] / 1000000.0  , 'children' : []}
			if budget['moyens_des_services'] != 0:
				obj['budget']['children']['gestion']['children'].append({ 'ref' : 'moyens_des_services' , 'montant' : budget['moyens_des_services'] / 1000000.0 })
			if budget['interventions_publiques'] != 0:
				obj['budget']['children']['gestion']['children'].append({ 'ref' : 'interventions_publiques' , 'montant' : budget['interventions_publiques'] / 1000000.0 })
			if budget['remunerations_publique'] != 0:
				obj['budget']['children']['gestion']['children'].append({ 'ref' : 'remunerations_publique' , 'montant' : budget['remunerations_publique'] / 1000000.0 })
			if budget['disposition_urgence'] != 0:
				obj['budget']['children']['gestion']['children'].append({ 'ref' : 'disposition_urgence' , 'montant' : budget['disposition_urgence'] / 1000000.0 })
			if budget['Avantages_dette_publique'] != 0:
				obj['budget']['children']['gestion']['children'].append({ 'ref' : 'Avantages_dette_publique' , 'montant' : budget['Avantages_dette_publique'] / 1000000.0 })
		if budget['fonds_speciaux'] != 0:
			obj['budget']['children']['fonds_speciaux'] =  { 'montant' : budget['fonds_speciaux'] / 1000000.0 , 'children' : [] }
		if ministere != None:
			obj['ministere'] = DBRef(collection='ministeres',id=ministere['_id'])
			if budget_type == "LF":
				prev = g.db.find_one('budgets' , {'ministere' : DBRef(collection='ministeres',id=ministere['_id']) , 'annee' : int(year)-1 , 'type' : 'LFC'}) or g.db.find_one('budgets' , {'ministere' : DBRef(collection='ministeres',id=ministere['_id']) , 'annee' : int(year)-1 , 'type' : 'LF'})
			elif budget_type == "LFC":
				prev = g.db.find_one('budgets' , {'ministere' : DBRef(collection='ministeres',id=ministere['_id']) , 'annee' : int(year) , 'type' : 'LF' })
				if prev == None:
					prev = g.db.find_one('budgets' , {'ministere' : DBRef(collection='ministeres',id=ministere['_id']) , 'annee' : int(year)-1 , 'type' : 'LFC'}) or g.db.find_one('budgets' , {'ministere' : DBRef(collection='ministeres',id=ministere['_id']) , 'annee' : int(year)-1 , 'type' : 'LF'})
			if prev == None:
				obj['budget']['evolution'] = 0
			else:
				obj['budget']['evolution'] = ( obj['budget']['total'] - prev['budget']['total'] ) / (prev['budget']['total'] / 1.0 )
			g.db.insert('budgets',obj)
		else:
			if budget['fr'] != 'Total':

				if budget['fr'] == "La Dette Publique" or budget['fr'] == "La Dette Publqiue":
					obj['budget']['titre'] = {'fr': u'Dette publique', 'ar': u'دين عمومي'}
					obj['budget']['icon']  = 'icon_dette.png'
					obj['budget']['id']	   = 'dette_publique'
				else:
					obj['budget']['titre'] = { 'ar' : budget['ar'] , 'fr' : budget['fr']}
					obj['budget']['id']	   = budget['fr'].replace(' ','_').replace(u'é','e')
				g.db.insert('budgets',obj)



@login_required
@editor_required
def add_budget():
	if request.method == 'POST':
		if 'budget_XL' not in request.files:
			return render_template('add_budget.html',error ='please select a file')
		elif request.form.get('budget_year') == '':
			return render_template('add_budget.html',error ='please type in the year')
		elif request.form.get('budget_type') not in ['LF' , 'LFC']:
			return render_template('add_budget.html',error ='please select a valid type ( LF or LFC)')
		else:
			xl = request.files['budget_XL']
			if xl.filename == '':
				return render_template('add_budget.html',error ='select a file please')
			else:
				if xl.filename[-4:] in ALLOWED_EXTENSIONS:
					xl.save(UPLOAD_FOLDER+xl.filename)
					link_budgets( budget_to_json(UPLOAD_FOLDER+xl.filename) , request.form.get('budget_year') , request.form.get('budget_type'))
					return redirect(url_for('admin.budgets'))
				else:
					return render_template('add_budget.html',error ='select a valid file please')
	return render_template('add_budget.html')

@login_required
@editor_required
def view_budgets():
	list_buds = []
	budgets = g.db.find('budgets',)
	for budget in budgets:
		list_buds.append({'year' : budget['annee'] , 'type' : budget['type']})
	list_buds = sorted(dict(((v['year'] , v['type']),v) for v in list_buds).values())
	return render_template('budgets.html' , budgets = list_buds)

@login_required
@editor_required
def delete_budget(year , budget_type):
	g.db.remove('budgets' , {'annee' : int(year) , 'type' : budget_type})
	return redirect(url_for('admin.budgets'))
